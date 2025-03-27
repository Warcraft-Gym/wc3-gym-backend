import logging
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from custom_exceptions import NotFoundException
from flasgger import swag_from
from src.dtos.season_dto import SeasonDTO
from src.dtos.user_dto import UserDTO
from src.dtos.team_dto import TeamDTO
from src.dtos.series_dto import SeriesDTO
from src.dtos.match_dto import MatchDTO
from src.dtos.map_dto import MapDTO
import pandas as pd
import io
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.util.import_util import ImportUtil


from src.util.query_util import QueryUtil

logger = logging.getLogger(__name__)


import_blueprint = Blueprint('import_api', __name__)

# import export endpoints
@import_blueprint.route('/import', methods=['POST'])
@jwt_required()
@swag_from({
    'summary': 'Import a google spreadsheet with the information for a GNL season',
    'description': 'Updates the database based on the import sheet',
    'tags': ['import export'],
    'security': [{'BearerAuth': []}],
    'parameters': [
        {'name': 'season_id', 'in': 'query', 'type': 'integer', 'required': False},
        {'name': 'season_name', 'in': 'query', 'type': 'string', 'required': False},
        {
            'name': 'file',
            'in': 'formData',
            'type': 'file',
            'required': True,
            'description': 'File to be uploaded (Google Sheet)'
        }
    ],
    'consumes': [
        'multipart/form-data'
    ],
    'responses': {
        200: {'description': 'Season updated successfully'},
        400: {
            'description': 'Bad Request',
            'examples': {
                'application/json': {
                    'error': 'No file part'
                }
            }
        },
        500: {'description': 'Internal server error'}
    }
})
def import_season():
    try:
        season_id = int(request.args.get('season_id')) if request.args.get('season_id') else None
        season_name = request.args.get('season_name')
        if 'file' not in request.files:
            return jsonify({"error": "No file part"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        player_name_id = {}
        if file and file.filename.endswith(('.xlsx', '.xls')):
            file_stream = io.BytesIO(file.read())
            
            # Load the Google Sheet into a DataFrame
            df_players = pd.read_excel(file_stream, sheet_name='Players')

            teams = []
            teams_players = {}
            for index, row in df_players.iterrows():
                if not ImportUtil.isNa(row['Bnet (no ID)']):
                    continue
                user_data = {
                        'name': ImportUtil.isNa(row['Bnet (no ID)']),
                        'battleTag': ImportUtil.isNa(row['Bnet']),
                        'discordTag': ImportUtil.isNa(row['Discord']),
                        'race': ImportUtil.getRaceEnumString(ImportUtil.isNa(row['Race'])),
                        'mmr': ImportUtil.isNa(row['MMR']),
                        'country':   ImportUtil.getCountryEnumString(ImportUtil.isNa(row['Country']))
                    }

                if not user_data.get('battleTag'):
                    raise Exception(f"User without BattleTag found: {user_data.get('name')}")
                query = QueryUtil.parseQuery("battleTag == " +  user_data.get('battleTag'))
                if not query or not query.elementA:
                    raise Exception(f"No valid query found: {"battleTag == " +  user_data.get('battleTag')}")
                users = import_blueprint.user_app_service.search(query)
                user = None
                if not users:
                    user = import_blueprint.user_app_service.create_user(UserDTO(user_data))
                else:
                    user = import_blueprint.user_app_service.update_user(users[0].id, UserDTO(user_data))
                player_name_id[user.name] = user.id

                team_name = ImportUtil.isNa(row['Team Abbr'])
                if team_name and not teams_players.get(team_name):
                    team_data = {
                        'name' : ImportUtil.isNa(row['Team Abbr'])
                    }
                    teams.append(team_data)
                    teams_players[team_name] = [user.id]
                elif team_name:
                    players = teams_players.get(team_name)
                    players.append(user.id)
                    
            excel_file = pd.ExcelFile(file_stream)
            sheet_names = excel_file.sheet_names
            week_sheets = [name for name in sheet_names if name.lower().startswith("week")]

            # Count the number of "Week" sheets
            number_of_week_sheets = len(week_sheets)

            season_data = {
                'name' : season_name,
                'number_weeks' : number_of_week_sheets
            }
            if season_id and season_name:
                import_blueprint.season_app_service.update_season(season_id, SeasonDTO(season_data))
            elif season_name:
                query = QueryUtil.parseQuery("name == " +  season_name)
                if not query or not query.elementA:
                    raise Exception(f"No valid query found: {"name == " +  season_name}")
                found_seasons = import_blueprint.season_app_service.search(query)
                if not found_seasons:
                    season_id = import_blueprint.season_app_service.create_season(SeasonDTO(season_data)).id
                else: 
                    season_id = found_seasons[0].id
                    import_blueprint.season_app_service.update_season(season_id, SeasonDTO(season_data))
            
            team_ids = []
            team_name_id = {}
            for team_data in teams:
                query = QueryUtil.parseQuery("name==" +  team_data.get('name'))
                if not query or not query.elementA:
                    raise Exception(f"No valid query found: {"name == " +  team_data.get('name')}")
                found_teams = import_blueprint.team_app_service.search(query)
                team = None
                if not found_teams:
                    team = import_blueprint.team_app_service.create_team(TeamDTO(team_data))
                else:
                    team = import_blueprint.team_app_service.update_team(found_teams[0].id, TeamDTO(team_data))
                team_name_id[team.name] = team.id
                team_ids.append(team.id)
                players = teams_players.get(team.name)
                import_blueprint.team_app_service.addPlayers(team.id, season_id, players)

            import_blueprint.season_app_service.addTeams(season_id, team_ids)

            for i in range(1,number_of_week_sheets+1):
                week = pd.read_excel(file_stream, sheet_name=f"Week {i}", header=None)
                matchups_rows = []
                date_frame = None
                fixed_map_short = None
                for index, row in week.iterrows():
                    if index==1:
                        date_frame = row[1]
                        fixed_map_short = row[3].split(":")[1].strip()
                    if "VS" in row.values:
                        matchups_rows.append(index)
                fixed_map = None
                if fixed_map_short:
                    q_string = f"shortname == {fixed_map_short}"
                    query = QueryUtil.parseQuery(q_string)
                    maps = import_blueprint.map_app_service.search(query)
                    if not maps:
                        map_data = {
                            "name": fixed_map_short,
                            "shortname": fixed_map_short
                        }
                        fixed_map = import_blueprint.map_app_service.create_map(MapDTO(map_data))
                    elif maps and len(maps) == 1:
                        fixed_map = maps[0]

                matchups_rows.append(len(week))
                for start, end in zip(matchups_rows, matchups_rows[1:]): 
                    team1 = None
                    team2 = None
                    match = None
                    for index, row in week.iloc[start:end].iterrows():
                        if index == start:
                            team1 = ImportUtil.isNa(row[0])
                            team2 = ImportUtil.isNa(row[2])
                            if not team1 or not team2:
                                raise Exception('Team 1 or Team 2 not properly identified in row:]')
                            team1_id = team_name_id[team1]
                            team2_id = team_name_id[team2]
                            q_string = f"team1_id=={team1_id} and team2_id=={team2_id} and season_id=={season_id}"
                            query = QueryUtil.parseQuery(q_string)
                            if not query or not query.elementA:
                                raise Exception(f"No valid query found: {q_string}")
                            found_matches = import_blueprint.match_app_service.search(query)
                            if not found_matches:
                                match_data = {
                                    'team1_id': team1_id,
                                    'team2_id': team2_id,
                                    'season_id': season_id,
                                    'playday': i,
                                    'fixed_map_id': fixed_map.id,
                                    'date_frame': date_frame
                                }
                                match = import_blueprint.match_app_service.create_match(MatchDTO(match_data))
                            else:
                                match = found_matches[0]
                            continue
                        if "Date" in row.values:
                            continue
                        if pd.isnull(row[5]) or pd.isnull(row[8]):
                            continue
                        if not team1 or not team2:
                            raise Exception(f"Teams could not be identified for Week {i} and row range {start}/{end}")
                        series_id = None
                        if ImportUtil.isNa(row[0]):
                            series_id = row[0]

                        player1_name=row[5]
                        player2_name=row[8]
                        

                        player1_id = player_name_id[player1_name.rstrip("*")]
                        player2_id = player_name_id[player2_name.rstrip("*")]
                        host_player_id = None
                        if player1_name.endswith("*"):
                            host_player_id = player1_id
                        elif player2_name.endswith("*"):
                            host_player_id = player2_id
                        date_time = None
                        if ImportUtil.isNa(row[4]) and ImportUtil.isNa(row[3]):
                            date_time = datetime.combine(row[4], row[3])
                        series_data = {
                            'caster': ImportUtil.isNa(row[1]),
                            'date_time': date_time,
                            'match_id' : match.id,
                            'player1_id': player1_id,
                            'player1_score': ImportUtil.isNa(row[6]),
                            'player2_score': ImportUtil.isNa(row[7]),
                            'player2_id': player2_id,
                            'host_player_id': host_player_id
                        }

                        series=None
                        if series_id:
                            series = import_blueprint.series_app_service.get_series(series_id)
                        else:
                            series_q_string = f"match_id=={match.id} and player1_id=={player1_id} and player2_id=={player2_id}"
                            series_query = QueryUtil.parseQuery(series_q_string)
                            if not query or not query.elementA:
                                raise Exception(f"No valid query found: {series_q_string}")
                            found_series = import_blueprint.series_app_service.search(series_query)
                            if found_series:
                                series = found_series[0]
                        if not series:
                            series = import_blueprint.series_app_service.create_series(SeriesDTO(series_data))
                        else:
                            series = import_blueprint.series_app_service.update_series(series.id, SeriesDTO(series_data))                       

            
            return jsonify({"message": "File uploaded successfully and data inserted into database"}), 200
        else:
            return jsonify({"error": "File type not allowed"}), 400
    except Exception as e:
        logger.error(e)
        return jsonify({"error": str(e)}), 500


@import_blueprint.route('/export', methods=['POST'])
@jwt_required()
@swag_from({
    'summary': 'Export a google spreadsheet with the information for a GNL season',
    'description': 'Export an exel sheet with the data of one season',
    'tags': ['import export'],
    'security': [{'BearerAuth': []}],
    'parameters': [
        {'name': 'season_id', 'in': 'query', 'type': 'integer', 'required': False},
        {'name': 'season_name', 'in': 'query', 'type': 'string', 'required': False}
    ],
    'responses': {
        200: {
            'description': 'A downloadable Excel file with user and team information',
            'content': {
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': {
                    'schema': {
                        'type': 'string',
                        'format': 'binary'
                    }
                }
            }
        },
        500: {'description': 'Internal server error'}
    }
})
def export_season():
    try:
        season_id = int(request.args.get('season_id')) if request.args.get('season_id') else None
        season_name = request.args.get('season_name')
        season_teams = []
        season = None
        if season_id:
            season = import_blueprint.season_app_service.get_season(season_id)
            if not season:
                raise NotFoundException(f"season not found by id: {season_id}")
        elif not season_id and season_name:
            query = QueryUtil.parseQuery(f"name == {season_name}")
            if not query or not query.elementA:
                raise Exception(f"No valid query found: name == {season_name}")
            found_seasons = import_blueprint.season_app_service.search(query)
            if not found_seasons:
                raise NotFoundException(f"season not found by name: {season_name}")
            else: 
                season = found_seasons[0]
                season_id = season.id


        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for week in range(1,season.number_weeks+1):

            # Create week worksheets
            week_sheet = workbook.create_sheet(title=f"Week {week}")
            for i in range(1,5):
                empty_cell = week_sheet.cell(row=1, column=i, value="")
                empty_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell = week_sheet.cell(row=1, column=5, value=f"WEEK {week}")
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Apply black background
            cell.font = Font(color="FFFFFF", bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")
            for i in range(6,9):
                empty_cell = week_sheet.cell(row=1, column=i, value="")
                empty_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            week_sheet.append([""])
            week_sheet.append([""])
            q_string = f"playday=={week} and season_id=={season.id}"
            query = QueryUtil.parseQuery(q_string)
            if not query or not query.elementA:
                raise Exception(f"No valid query found: {q_string}")
            matches = import_blueprint.match_app_service.search(query)
            if matches:
                for match in matches:
                    week_sheet.append([match.team1.name,"VS",match.team2.name])
                    q_string = f"match_id=={match.id}"
                    query = QueryUtil.parseQuery(q_string)
                    if not query or not query.elementA:
                        raise Exception(f"No valid query found: {q_string}")
                    series_list = import_blueprint.series_app_service.search(query)
                    week_sheet.append(["","Caster Twitch Handle","","Time EDT (US East)","Date", match.team1.name,"Score 1","Score 2",match.team2.name,"TeamScore 1","TeamScore 2"])
                    for series in series_list:
                        date = None
                        time = None
                        if series.date_time:
                            date = series.date_time.date()
                            time = series.date_time.time()
                        week_sheet.append([series.id,series.caster,"",time,date,series.player1.name,series.player1_score,series.player2_score,series.player2.name])
                    week_sheet.append([""])
                    week_sheet.append([""])

        # Create worksheets
        ranking_sheet = workbook.create_sheet(title='Ranking')
        user_sheet = workbook.create_sheet(title='Players')
        user_sheet.append(['Bnet', 'Bnet (no ID)', 'Bnet + Host', 'Discord', 'Race', 'Team Abbr', 'MMR', 'Country'])
        player_team_sheet = workbook.create_sheet(title='Player Team Assignment')
        player_team_sheet.append(['battle net name', 'team'])
        player_website_sheet = workbook.create_sheet(title='Player Website')
        player_website_sheet.append(['player', 'race_image', 'w3_champions_profile', 'mmr', 'country', 'team', 'captain'])
        ranking_sheet.append([f"{season.name} Rankings"])
        ranking_sheet.append([""])
        ranking_header = []
        ranking_header.append('Rank')
        ranking_header.append('Team')
        for number in range(season.number_weeks):
            ranking_header.append(f"Week {number}")
        ranking_header.append('Points')
        ranking_header.append('Points Against (PA)')
        ranking_header.append('Points Available')
        ranking_sheet.append(ranking_header)
        rank = 1

        season_teams = import_blueprint.team_app_service.get_teams_season(season_id)
        for team in season_teams:
            # ranking sheet
            team_rank = []
            team_rank.append(rank)
            team_rank.append(team.name)
            for number in range(season.number_weeks):
                team_rank.append(0)
            team_rank.append(team.seasons_info[0].final_score)
            team_rank.append(team.seasons_info[0].points_available)
            team_rank.append(team.seasons_info[0].points_against)
            ranking_sheet.append(team_rank)
            rank += 1
            # Player sheet
            players = team.player_by_season[season_id]
            for user in players:
                player_team_sheet.append([user.battleTag,team.name])
                player_website_sheet.append([user.name, ImportUtil.getRaceImage(user.race),ImportUtil.getW3ChampionURL(user.battleTag),user.mmr,ImportUtil.getCountryNameString(user.country),team.name])
                user_sheet.append([user.battleTag, user.name,f"{user.name}*",user.discordTag,ImportUtil.getRaceNameString(user.race),team.name, user.mmr,ImportUtil.getCountryNameString(user.country)])
                
        excel_stream = BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)

        # Return the Excel file for download
        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=f'{season.name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(e)
        return jsonify({"error": str(e)}), 500